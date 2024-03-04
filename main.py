import os
from openpyxl import load_workbook
from pyautogui import click, write, press, doubleClick
from time import sleep
import tkinter as tk

def btn(): 

    campoCodigoInterno = ''
    campoEscreverEstoque = ''
    btnAlterar = ''

    # Função para ler as configurações do arquivo
    def ler_configuracoes():
        configuracoes = {}
        with open("config.txt", "r") as arquivo:
            for linha in arquivo:
                chave, valor = linha.strip().split("=")
                configuracoes[chave.strip()] = tuple(map(int, valor.strip().split(",")))
        return configuracoes

    # Obtendo as configurações
    configuracoes = ler_configuracoes()

    # Atribuindo os valores às variáveis
    campoCodigoInterno = configuracoes["campoCodigoInterno"]
    campoEscreverEstoque = configuracoes["campoEscreverEstoque"]
    btnAlterar = configuracoes["btnAlterar"]

    sleep(5)
    def processar_coluna(nome_arquivo, nome_coluna):
        caminho_arquivo = os.path.join("relatorio", nome_arquivo)
        if not os.path.exists(caminho_arquivo):
            print(f"O arquivo '{nome_arquivo}' não foi encontrado.")
            return

        try:
            workbook = load_workbook(caminho_arquivo)
            if "Planilha1" in workbook.sheetnames:  
                planilha = workbook["Planilha1"]  
                coluna_index_codigo = None
                coluna_index_estoque = None
                for col in range(1, planilha.max_column + 1):
                    if planilha.cell(row=1, column=col).value == "codigo":
                        coluna_index_codigo = col
                    elif planilha.cell(row=1, column=col).value == "estoque":
                        coluna_index_estoque = col
                    if coluna_index_codigo is not None and coluna_index_estoque is not None:
                        break

                if coluna_index_codigo is not None and coluna_index_estoque is not None:
                    for row in range(2, planilha.max_row + 1):
                        codigo = planilha.cell(row=row, column=coluna_index_codigo).value
                        estoque = planilha.cell(row=row, column=coluna_index_estoque).value
                        if codigo is not None:
                            with open("codigos_alterados.txt", "a") as arquivo_codigos:
                                arquivo_codigos.write(f"{codigo}\n")
                            
                            sleep(0.5)
                            # Clique no campo onde você quer escrever o código
                            doubleClick(campoCodigoInterno)
                            # Escreva o código atual
                            write(str(codigo))
                            # Pressione Enter para confirmar
                            sleep(0.1)
                            press("enter")
                            press("enter")
                            # Clique no campo onde você quer escrever o estoque
                            doubleClick(campoEscreverEstoque)
                            # Escreva o valor da coluna estoque dessa repetição
                            write(str(estoque))
                            press("enter")
                            sleep(0.2)
                            click(btnAlterar)
                            press("enter")
                            press("enter")
                else:
                    print("As colunas 'codigo' e 'estoque' não foram encontradas na planilha.")
            else:
                print("A planilha 'relatorio' não foi encontrada no arquivo.")
        except Exception as e:
            print(f"Ocorreu um erro ao abrir o arquivo '{nome_arquivo}': {e}")

    processar_coluna("relatorio.xlsx", "codigo")

# Criar a janela
janela = tk.Tk()

janela.geometry('400x200')
janela.title('Alterar Estoque Fiscal By Fábio')
# Adicionar um rótulo à janela
label = tk.Label(janela, text="Clique no botão para iniciar o bot")
label.pack(padx=0, pady=40)  # pady é o preenchimento vertical

# Adicionar um botão à janela
button = tk.Button(janela, text="Iniciar", command=btn, width=30, height=10, bg='blue')
button.pack(pady=30)  # pady é o preenchimento vertical

# Iniciar o loop de eventos da janela
janela.mainloop()